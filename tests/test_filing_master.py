"""Tests for the per-period filing-returns master + split."""

from dataclasses import replace

from openpyxl import load_workbook

from nport.config import parse_filing
from nport.custodian import CustodianRow
from nport.filing_master import (
    HEADER,
    RISK_HEADER,
    _aggregate_risk,
    _clean_return,
    _is_investment_grade,
    _maturity_bucket,
    _month_ranges,
    _return_formula,
    _signed_date,
    build_filing_master,
    read_filing_master,
    read_risk_sheet,
    split_filing_master,
)


def _crow(account, market_value, net="1000000", cusip="000000000", security_name=None):
    return CustodianRow(
        date="06/01/2026", account=account, stock_ticker=account, cusip=cusip,
        security_name=security_name or f"{account} ETF", shares="100", price="10",
        market_value=market_value, weightings="1%", net_assets=net,
        shares_outstanding="1000", creation_units="1", money_market_flag="",
    )


def _treasury_crow(account, market_value, cusip="912797UL9"):
    return _crow(account, market_value, cusip=cusip,
                 security_name="United States Treasury Bill 0% 10/22/2026")


def _corp_crow(account, market_value, cusip="00081TAK4"):
    return _crow(account, market_value, cusip=cusip,
                 security_name="ACCO Brands Corp 4.250% 03/15/2029")


def test_month_ranges():
    assert _month_ranges("2026-06") == [
        ("20260401", "20260430"), ("20260501", "20260531"), ("20260601", "20260630")]
    # crosses the year boundary
    assert _month_ranges("2026-01") == [
        ("20251101", "20251130"), ("20251201", "20251231"), ("20260101", "20260131")]


def test_return_formula():
    f = _return_formula("$B2", "20260401", "20260430")
    assert f.startswith('=BDP($B2,"CUST_TRR_RETURN_HOLDING_PER"')
    assert '"CUST_TRR_START_DT","20260401"' in f
    assert '"CUST_TRR_END_DT","20260430"' in f
    assert '"CUST_TRR_CRNCY","USD"' in f


def test_signed_date():
    assert _signed_date("2026-06") == "2026-07-31"
    assert _signed_date("2026-12") == "2027-01-31"


def test_clean_return():
    assert _clean_return("11.32327") == "11.32"   # rounded to 2dp
    assert _clean_return("-0.5") == "-0.50"
    assert _clean_return("") == "N/A"             # uncalculated
    assert _clean_return("#N/A N/A") == "N/A"     # Bloomberg error → valid sentinel


def test_header_excludes_risk_json():
    for k in ("curMetricsJson", "creditSprdRiskIgJson", "creditSprdRiskNonigJson"):
        assert k not in HEADER
    assert HEADER[:2] == ["Account", "bbgid"]


def test_build_read_roundtrip(tmp_path):
    rows = [_crow("AAA", "-5000", net="2000000"), _crow("BBB", "1000")]
    p = tmp_path / "fm.xlsx"
    assert build_filing_master(rows, "2026-06", p) == 2
    recs = {r["Account"]: r for r in read_filing_master(p)}
    a = recs["AAA"]
    assert a["netAssets"] == "2000000.00"
    assert a["totLiabs"] == "5000.00"            # abs(negative market value)
    assert a["totAssets"] == "2005000.00"        # net + liabs
    assert a["rtn1"] == "N/A" and a["rtn3"] == "N/A"   # uncalculated off-terminal
    assert a["submissionType"] == "NPORT-P" and a["repPdEnd"] == "2026-06-30"
    assert a["mon1Sales"] == "0" and a["isNonCashCollateral"] == "N"


def test_split_produces_parseable_filing(tmp_path):
    rows = [_crow("AAA", "-5000")]
    p = tmp_path / "fm.xlsx"
    build_filing_master(rows, "2026-06", p)
    funds = tmp_path / "funds"
    res = split_filing_master(p, funds, "2026-06")
    assert len(res) == 1
    fpath = funds / "aaa" / "filings" / "2026-06" / "filing_data.txt"
    assert fpath.is_file()
    fd = parse_filing(fpath)            # must parse into FilingData with no missing keys
    assert fd.submission_type == "NPORT-P"
    assert fd.rep_pd_end == "2026-06-30"
    assert fd.rtn1 == "N/A"
    assert fd.net_assets == "1000000.00"
    assert fd.mon1_sales == "0"


# ── Capital flows (AP order book) ─────────────────────────────


def test_build_applies_ap_order_flows(tmp_path):
    orders = tmp_path / "orders.csv"
    orders.write_text(
        "Ticker,Side,Trade Date,Notional,Status\n"
        "AAA,CREATE,4/10/2026,1000,ACCEPTED\n"
        "AAA,REDEEM,6/2/2026,250,ACCEPTED\n"
        "AAA,CREATE,6/3/2026,40,CANCELLED\n",       # excluded
        encoding="utf-8",
    )
    p = tmp_path / "fm.xlsx"
    build_filing_master([_crow("AAA", "1000")], "2026-06", p, ap_orders_path=orders)
    rec = {r["Account"]: r for r in read_filing_master(p)}["AAA"]
    assert rec["mon1Sales"] == "1000.00"        # April create
    assert rec["mon3Redemption"] == "250.00"    # June redeem
    assert rec["mon3Sales"] == "0.00"           # cancelled order excluded
    assert rec["mon1Reinvestment"] == "0"       # never sourced


# ── B.3 risk metrics ──────────────────────────────────────────


def test_maturity_bucket():
    assert _maturity_bucket(0.1) == "3month"
    assert _maturity_bucket(0.8) == "1year"
    assert _maturity_bucket(4.0) == "5year"
    assert _maturity_bucket(9.0) == "10year"
    assert _maturity_bucket(25.0) == "30year"


def test_is_investment_grade():
    assert _is_investment_grade("AAA") and _is_investment_grade("BBB-")
    assert _is_investment_grade("") and _is_investment_grade("#N/A")   # treasuries / unrated
    assert not _is_investment_grade("BB+") and not _is_investment_grade("B")


def test_aggregate_risk_shapes_and_values():
    import json
    rows = [
        {"durAdj": "2.5", "spreadDur": "2.4", "maturity": "2031-06-30",
         "valUSD": "1000000", "ratingSP": "BBB"},     # 5yr, IG
        {"durAdj": "0.3", "spreadDur": "", "maturity": "2026-09-30",
         "valUSD": "500000", "ratingSP": ""},         # 3mo, treasury (IG, spread→dur)
        {"durAdj": "4.0", "spreadDur": "3.9", "maturity": "2032-06-30",
         "valUSD": "200000", "ratingSP": "BB+"},      # 5yr, non-IG
    ]
    cur, ig, nonig = _aggregate_risk(rows, "2026-06")
    cm = json.loads(cur)[0]
    assert cm["curCd"] == "USD"
    assert cm["dv01_5year"] == f"{2.5*1e6*1e-4 + 4.0*2e5*1e-4:.2f}"   # 250 + 80
    assert cm["dv01_3month"] == f"{0.3*5e5*1e-4:.2f}"                  # 15
    assert cm["dv100_5year"] == f"{2.5*1e6*1e-2 + 4.0*2e5*1e-2:.2f}"
    ig_d, nonig_d = json.loads(ig), json.loads(nonig)
    assert ig_d["5year"] == f"{2.4*1e6*1e-4:.2f}"                      # BBB only
    assert ig_d["3month"] == f"{0.3*5e5*1e-4:.2f}"                     # treasury spread→dur
    assert nonig_d["5year"] == f"{3.9*2e5*1e-4:.2f}"                   # BB+ bond


def test_aggregate_risk_empty_off_terminal():
    # uncalculated formulas read back as blank / #N/A → no B.3 emitted
    rows = [{"durAdj": "", "spreadDur": "", "maturity": "", "valUSD": "1000", "ratingSP": ""}]
    assert _aggregate_risk(rows, "2026-06") == ("", "", "")


def test_build_writes_risk_sheet(tmp_path):
    p = tmp_path / "fm.xlsx"
    rows = [_crow("EQ", "1000"), _treasury_crow("CBIL", "5000"), _corp_crow("CHYG", "3000")]
    build_filing_master(rows, "2026-06", p)
    wb = load_workbook(p)                       # formulas (not data_only)
    assert "risk" in wb.sheetnames
    ws = wb["risk"]
    body = [[c.value for c in row] for row in ws.iter_rows(min_row=2)]
    assert len(body) == 2                        # equity excluded
    by_acct = {r[0]: r for r in body}
    assert by_acct["CBIL"][RISK_HEADER.index("bbgid")] == "912797UL9 Govt"
    assert by_acct["CHYG"][RISK_HEADER.index("bbgid")] == "00081TAK4 Corp"
    dur = by_acct["CBIL"][RISK_HEADER.index("durAdj")]
    assert dur.startswith('=BDP(') and "DUR_ADJ_MID" in dur


def test_split_b3_round_trips_to_valid_xml(tmp_path, sample_data):
    """A calculated risk sheet → curMetricsJson that builds a valid B.3 block."""
    from lxml import etree

    from nport.builder import NportBuilder
    from nport.constants import NS_NPORT

    # Simulate a post-terminal workbook: write literal (calculated) durations.
    p = tmp_path / "fm.xlsx"
    build_filing_master([_corp_crow("CHYG", "1000000")], "2026-06", p)
    wb = load_workbook(p)
    ws = wb["risk"]
    hdr = {name: i for i, name in enumerate(RISK_HEADER)}
    ws.cell(row=2, column=hdr["durAdj"] + 1, value="2.47")
    ws.cell(row=2, column=hdr["spreadDur"] + 1, value="2.47")
    ws.cell(row=2, column=hdr["maturity"] + 1, value="2029-03-15")
    ws.cell(row=2, column=hdr["ratingSP"] + 1, value="B")
    wb.save(p)

    risk_rows = read_risk_sheet(p)
    cur, ig, nonig = _aggregate_risk(risk_rows, "2026-06")
    assert cur and "USD" in cur

    config, filing, holdings = sample_data
    # dataclasses.replace → fresh FilingData; never mutate the session-scoped fixture.
    filing = replace(filing, cur_metrics_json=cur,
                      credit_sprd_risk_ig_json=ig, credit_sprd_risk_nonig_json=nonig)
    root = etree.fromstring(NportBuilder(config, filing, holdings).to_xml_bytes())
    ns = {"n": NS_NPORT}
    assert root.find(".//n:curMetrics/n:curMetric/n:curCd", ns).text == "USD"
    assert root.find(".//n:curMetrics/n:curMetric/n:intrstRtRiskdv01", ns) is not None
    # B (junk) bond → spread DV01 lands in the non-IG element
    nonig_el = root.find(".//n:creditSprdRiskNonInvstGrade", ns)
    assert nonig_el is not None and float(nonig_el.get("period5Yr")) > 0


def test_split_off_terminal_omits_b3(tmp_path):
    """Uncalculated risk formulas → no B.3 keys in filing_data.txt (graceful)."""
    p = tmp_path / "fm.xlsx"
    build_filing_master([_corp_crow("CHYG", "1000000")], "2026-06", p)
    funds = tmp_path / "funds"
    split_filing_master(p, funds, "2026-06")
    fd = parse_filing(funds / "chyg" / "filings" / "2026-06" / "filing_data.txt")
    assert fd.cur_metrics_json == ""
