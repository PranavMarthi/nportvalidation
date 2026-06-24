"""Tests for the global master spreadsheet (refresh / split / seed)."""

import csv
from pathlib import Path

from openpyxl import load_workbook

from nport.custodian import EQUITY_HEADERS, parse_custodian_csv
from nport.master_sheet import (
    MASTER_ENRICHMENT_COLUMNS,
    _BBG_FORMULA_COLUMNS,
    _bbg_spec_key,
    _normalize_coupon_kind,
    apply_bloomberg_formulas,
    read_master_xlsx,
    refresh_master,
    seed_master_from_per_fund,
    split_master,
    write_master_xlsx,
)


# ── Helpers ───────────────────────────────────────────────────

_CUSTODIAN_HEADERS = [
    "Date", "Account", "StockTicker", "CUSIP", "SecurityName",
    "Shares", "Price", "MarketValue", "Weightings", "NetAssets",
    "SharesOutstanding", "CreationUnits", "MoneyMarketFlag",
]


def _eq(account, ticker, cusip, name, mmflag=""):
    return {
        "Date": "06/01/2026", "Account": account, "StockTicker": ticker,
        "CUSIP": cusip, "SecurityName": name, "Shares": "100",
        "Price": "10", "MarketValue": "1000", "Weightings": "1.00%",
        "NetAssets": "100000", "SharesOutstanding": "5000",
        "CreationUnits": "10", "MoneyMarketFlag": mmflag,
    }


def _write_custodian(tmp_path, rows) -> Path:
    p = tmp_path / "2026-06_holdings.csv"
    with open(p, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=_CUSTODIAN_HEADERS)
        w.writeheader()
        w.writerows(rows)
    return p


def _write_per_fund_csv(funds_dir, fund, header, rows) -> Path:
    d = funds_dir / fund
    d.mkdir(parents=True, exist_ok=True)
    p = d / "security_master.csv"
    with open(p, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=header)
        w.writeheader()
        w.writerows(rows)
    return p


def _read_csv(path):
    with open(path, newline="", encoding="utf-8") as f:
        r = csv.DictReader(f)
        header = [h.strip() for h in (r.fieldnames or [])]
        rows = [{(k.strip() if k else k): (v or "") for k, v in row.items()} for row in r]
    return header, rows


# ── refresh_master ────────────────────────────────────────────


def test_refresh_adds_new_holdings(tmp_path):
    cust = _write_custodian(tmp_path, [
        _eq("FDRS", "ABNB", "009066101", "Airbnb Inc"),
        _eq("FDRS", "NVDA", "67066G104", "NVIDIA Corp"),
    ])
    master = tmp_path / "master.xlsx"
    stats = refresh_master(parse_custodian_csv(cust), master)
    assert stats["added"] == 2
    assert stats["kept"] == 0 and stats["removed"] == 0

    rows, _ = read_master_xlsx(master)
    tickers = {r["ticker"] for r in rows}
    assert tickers == {"ABNB", "NVDA"}
    assert all(r["Account"] == "FDRS" for r in rows)


def test_refresh_preserves_manual_edits(tmp_path):
    # Operator-owned (non-Bloomberg) fields survive a refresh. Bloomberg fields
    # are always formulas, so the durable manual data is on derivatives — here an
    # option's delta (options have no Bloomberg spec).
    cust = _write_custodian(tmp_path, [_opt_custodian("BUF", "SPY 12/18/2026 4800 C")])
    master = tmp_path / "master.xlsx"
    refresh_master(parse_custodian_csv(cust), master)

    rows, header = read_master_xlsx(master)
    rows[0]["delta"] = "0.61"
    write_master_xlsx(rows, header, master)

    # Refresh again with the same custodian — manual delta must survive.
    stats = refresh_master(parse_custodian_csv(cust), master)
    assert stats["kept"] == 1 and stats["added"] == 0
    rows2, _ = read_master_xlsx(master)
    assert rows2[0]["delta"] == "0.61"


def test_refresh_drops_removed_holdings(tmp_path):
    master = tmp_path / "master.xlsx"
    refresh_master(parse_custodian_csv(_write_custodian(tmp_path, [
        _eq("FDRS", "ABNB", "009066101", "Airbnb Inc"),
        _eq("FDRS", "NVDA", "67066G104", "NVIDIA Corp"),
    ])), master)

    stats = refresh_master(parse_custodian_csv(_write_custodian(tmp_path, [
        _eq("FDRS", "ABNB", "009066101", "Airbnb Inc"),
    ])), master)
    assert stats["removed"] == 1 and stats["kept"] == 1
    rows, _ = read_master_xlsx(master)
    assert {r["ticker"] for r in rows} == {"ABNB"}


def test_same_security_two_funds_stays_two_rows(tmp_path):
    cust = _write_custodian(tmp_path, [
        _eq("FDRS", "NVDA", "67066G104", "NVIDIA Corp"),
        _eq("BAY", "NVDA", "67066G104", "NVIDIA Corp"),
    ])
    master = tmp_path / "master.xlsx"
    refresh_master(parse_custodian_csv(cust), master)
    rows, _ = read_master_xlsx(master)
    accounts = sorted(r["Account"] for r in rows if r["ticker"] == "NVDA")
    assert accounts == ["BAY", "FDRS"]


def test_refresh_always_mirrors_full_custodian(tmp_path):
    # The master is strictly 1:1 with the custodian (sheet 2 row i ↔ sheet 1 row
    # i), so build-master rebuilds the WHOLE master from the custodian — the
    # accounts hint no longer carves out a partial master.
    cust = _write_custodian(tmp_path, [
        _eq("FDRS", "NVDA", "67066G104", "NVIDIA Corp"),
        _eq("BAY", "ABNB", "009066101", "Airbnb Inc"),
    ])
    master = tmp_path / "master.xlsx"
    refresh_master(parse_custodian_csv(cust), master, accounts=["FDRS"])
    rows, _ = read_master_xlsx(master)
    assert {r["Account"] for r in rows} == {"FDRS", "BAY"}


def test_refresh_idempotent(tmp_path):
    cust = _write_custodian(tmp_path, [
        _eq("FDRS", "ABNB", "009066101", "Airbnb Inc"),
        _eq("FDRS", "NVDA", "67066G104", "NVIDIA Corp"),
    ])
    master = tmp_path / "master.xlsx"
    refresh_master(parse_custodian_csv(cust), master)
    first = master.read_bytes()
    refresh_master(parse_custodian_csv(cust), master)
    # Re-run keeps everything; row set + values identical.
    rows_a, hdr_a = read_master_xlsx(master)
    assert len(rows_a) == 2
    # idempotent at the data level (byte-identical not guaranteed by openpyxl metadata)
    refresh_master(parse_custodian_csv(cust), master)
    rows_b, hdr_b = read_master_xlsx(master)
    assert hdr_a == hdr_b
    assert [r["ticker"] for r in rows_a] == [r["ticker"] for r in rows_b]


# ── Text format prevents corruption ───────────────────────────


def test_identifier_columns_stored_as_text(tmp_path):
    master = tmp_path / "master.xlsx"
    header = ["Account", "StockTicker", "CUSIP"] + list(MASTER_ENRICHMENT_COLUMNS)
    row = {c: "" for c in header}
    row.update({"Account": "FDRS", "StockTicker": "AIR", "CUSIP": "000361105",
                "ticker": "AIR", "cusip": "000361105", "name": "AAR Corp"})
    row2 = {c: "" for c in header}
    row2.update({"Account": "FDRS", "StockTicker": "RTX", "CUSIP": "75513E101",
                 "ticker": "RTX", "cusip": "75513E101", "name": "RTX Corp"})
    write_master_xlsx([row, row2], header, master)

    wb = load_workbook(master)
    ws = wb.active
    hdr = [c.value for c in ws[1]]
    cusip_col = hdr.index("cusip") + 1
    # Every data cell in the cusip column is Text formatted.
    for r in range(2, ws.max_row + 1):
        assert ws.cell(row=r, column=cusip_col).number_format == "@"

    # Values read back exactly — no 361105, no 7.55E+105.
    rows, _ = read_master_xlsx(master)
    cusips = {r["cusip"] for r in rows}
    assert cusips == {"000361105", "75513E101"}


# ── split_master ──────────────────────────────────────────────


def test_split_shape_equity_only(tmp_path):
    cust = _write_custodian(tmp_path, [_eq("FDRS", "ABNB", "009066101", "Airbnb Inc")])
    master = tmp_path / "master.xlsx"
    refresh_master(parse_custodian_csv(cust), master)
    funds = tmp_path / "funds"
    results = split_master(master, funds)
    assert len(results) == 1
    header, rows = _read_csv(results[0][1])
    assert header == list(EQUITY_HEADERS)  # exactly 9 cols
    assert len(rows) == 1 and rows[0]["ticker"] == "ABNB"


def test_split_dry_run_writes_nothing(tmp_path):
    cust = _write_custodian(tmp_path, [_eq("FDRS", "ABNB", "009066101", "Airbnb Inc")])
    master = tmp_path / "master.xlsx"
    refresh_master(parse_custodian_csv(cust), master)
    funds = tmp_path / "funds"
    results = split_master(master, funds, dry_run=True)
    assert results[0][2] == 1
    assert not (funds / "fdrs" / "security_master.csv").exists()


# ── seed migration round-trip ─────────────────────────────────


def _opt_custodian(account, name, shares="10"):
    return {"Date": "06/01/2026", "Account": account, "StockTicker": name,
            "CUSIP": "N/A", "SecurityName": name, "Shares": shares, "Price": "5",
            "MarketValue": "500", "Weightings": "0.50%", "NetAssets": "100000",
            "SharesOutstanding": "5000", "CreationUnits": "10", "MoneyMarketFlag": ""}


def test_seed_then_split_reproduces_per_fund(tmp_path):
    funds = tmp_path / "funds"
    # Equity fund — two equities (to check custodian order is preserved).
    _write_per_fund_csv(funds, "fdrs", list(EQUITY_HEADERS), [
        {"name": "Airbnb Inc", "lei": "549300HMUDNO0RY56D37", "title": "Airbnb Inc",
         "cusip": "009066101", "isin": "US0090661010", "ticker": "ABNB",
         "invCountry": "US", "assetCat": "EC", "issuerCat": "CORP"},
        {"name": "NVIDIA Corp", "lei": "549300S4KLFTLO7GSQ80", "title": "NVIDIA Corp",
         "cusip": "67066G104", "isin": "US67066G1040", "ticker": "NVDA",
         "invCountry": "US", "assetCat": "EC", "issuerCat": "CORP"},
    ])
    # Option fund with a manual delta. Ticker must equal the custodian-generated
    # option id: "SPX 12/18/2026 4800 C" -> "SPX-C4800-20261218".
    from nport.custodian import OPTION_HEADERS
    opt_row = {c: "" for c in OPTION_HEADERS}
    opt_row.update({"name": "SPX 12/18/2026 4800 C", "title": "SPX 12/18/2026 4800 C",
                    "cusip": "N/A", "ticker": "SPX-C4800-20261218", "invCountry": "US",
                    "assetCat": "DE", "issuerCat": "CORP", "derivCat": "OPT",
                    "counterpartyName": "Goldman Sachs International", "delta": "0.72",
                    "putOrCall": "Call", "writtenOrPur": "Purchased",
                    "exercisePrice": "4800", "exercisePriceCurCd": "USD",
                    "expDt": "2026-12-18", "refInstType": "indexBasket",
                    "refIndexName": "S&P 500 Index", "refIndexIdentifier": "SPX"})
    _write_per_fund_csv(funds, "buf", list(OPTION_HEADERS), [opt_row])

    # Custodian defines the order: NVDA before ABNB in FDRS, then the option.
    cust = _write_custodian(tmp_path, [
        _eq("FDRS", "NVDA", "67066G104", "NVIDIA Corp"),
        _eq("FDRS", "ABNB", "009066101", "Airbnb Inc"),
        _opt_custodian("BUF", "SPX 12/18/2026 4800 C"),
    ])

    master = tmp_path / "master.xlsx"
    stats = seed_master_from_per_fund(funds, cust, master)
    assert stats["funds"] == 2 and stats["rows"] == 3

    # Master row order matches the custodian (NVDA, ABNB, option) — not sorted.
    mrows, _ = read_master_xlsx(master)
    assert [r["ticker"] for r in mrows] == ["NVDA", "ABNB", "SPX-C4800-20261218"]

    # Split reproduces per-fund content for every NON-Bloomberg column. The
    # Bloomberg-owned fields (isin/lei/invCountry/...) are now formulas resolved
    # on a terminal, so off-Bloomberg they read back empty (no fallback) — they
    # are excluded from the round-trip equality and checked separately below.
    out = tmp_path / "out"
    split_master(master, out)
    for fund in ("fdrs", "buf"):
        oh, orows = _read_csv(funds / fund / "security_master.csv")
        nh, nrows = _read_csv(out / fund / "security_master.csv")
        assert set(oh) == set(nh), fund
        common = [c for c in oh if c in nh and c not in _BBG_FORMULA_COLUMNS]
        o_by = {r["ticker"]: r for r in orows}
        n_by = {r["ticker"]: r for r in nrows}
        for t in o_by:
            assert {c: o_by[t][c] for c in common} == {c: n_by[t][c] for c in common}, (fund, t)

    # The manual (non-Bloomberg) delta + counterparty survived the round-trip.
    _, buf_rows = _read_csv(out / "buf" / "security_master.csv")
    assert buf_rows[0]["delta"] == "0.72"
    assert buf_rows[0]["counterpartyName"] == "Goldman Sachs International"
    # Equity Bloomberg fields are emptied until resolved on a terminal.
    _, fdrs_rows = _read_csv(out / "fdrs" / "security_master.csv")
    assert all(r["lei"] == "" and r["isin"] == "" and r["invCountry"] == "" for r in fdrs_rows)


# ── Bloomberg BDP formulas ────────────────────────────────────


def _equity(**over):
    row = {"Account": "FDRS", "cusip": "037833100",
           "ticker": "AAPL", "name": "Apple Inc", "title": "Apple Inc",
           "assetCat": "EC", "derivCat": "",
           "lei": "549300...", "isin": "US0378331005", "invCountry": "US"}
    row.update(over)
    return row


def _equity_master(rows, tmp_path):
    """Write equity rows to a master xlsx and return the raw worksheet."""
    from openpyxl import load_workbook
    from nport.master_sheet import IDENTITY_COLUMNS, MASTER_ENRICHMENT_COLUMNS
    header = IDENTITY_COLUMNS + list(MASTER_ENRICHMENT_COLUMNS)
    p = tmp_path / "m.xlsx"
    write_master_xlsx(rows, header, p)
    return load_workbook(p).active


def test_equity_gets_bbgid_and_simple_formulas(tmp_path):
    rows = [_equity(lei="N/A", isin="", invCountry="US")]
    n = apply_bloomberg_formulas(rows)
    assert n == 3  # isin, lei, invCountry (NOT cusip)
    # apply only sets the bbgid lookup key; the formulas are emitted on write.
    assert rows[0]["bbgid"] == "AAPL US Equity"

    ws = _equity_master(rows, tmp_path)
    hdr = [c.value for c in ws[1]]
    idx = {c: i + 1 for i, c in enumerate(hdr)}
    bb = idx["bbgid"]
    from openpyxl.utils import get_column_letter
    bb_ref = f"${get_column_letter(bb)}2"
    assert ws.cell(row=2, column=bb).value == "AAPL US Equity"
    # isin/lei/invCountry are BARE formulas referencing the bbgid CELL — no
    # IFERROR / embedded fallback.
    le = ws.cell(row=2, column=idx["lei"]).value
    assert le == f'=BDP({bb_ref},"LEGAL_ENTITY_IDENTIFIER")'
    assert "IFERROR" not in le
    assert ws.cell(row=2, column=idx["isin"]).value == f'=BDP({bb_ref},"ID_ISIN")'
    assert ws.cell(row=2, column=idx["invCountry"]).value == f'=BDP({bb_ref},"CNTRY_OF_DOMICILE")'
    # cusip is NOT a formula — it's the custodian literal.
    assert ws.cell(row=2, column=idx["cusip"]).value == "037833100"
    # name/title/ticker stay literal too.
    for f in ("name", "title", "ticker"):
        assert not str(ws.cell(row=2, column=idx[f]).value).startswith("=")


def test_cusip_literal_us_kept_foreign_na(tmp_path):
    # cusip is a literal from the row (custodian), never a formula.
    rows = [_equity(ticker="AMD", cusip="007903107", lei="N/A", isin="", invCountry="US"),
            _equity(ticker="AER", cusip="N/A", lei="549300SZYINBBLJQU475",
                    isin="NL0000687663", invCountry="NL", name="AerCap", title="AerCap")]
    apply_bloomberg_formulas(rows)
    master = tmp_path / "m.xlsx"
    from nport.master_sheet import IDENTITY_COLUMNS, MASTER_ENRICHMENT_COLUMNS
    write_master_xlsx(rows, IDENTITY_COLUMNS + list(MASTER_ENRICHMENT_COLUMNS), master)
    back, _ = read_master_xlsx(master)
    amd = next(r for r in back if r["ticker"] == "AMD")
    aer = next(r for r in back if r["ticker"] == "AER")
    assert amd["cusip"] == "007903107"        # US CUSIP literal
    assert aer["cusip"] == "N/A"              # foreign ordinary -> N/A
    # lei/country are bare formulas with NO fallback: uncalculated (no Bloomberg
    # terminal in the test) they read back EMPTY, so a failed lookup stays visible.
    assert aer["lei"] == "" and aer["invCountry"] == ""


def test_option_no_spec_treasury_isin_only(tmp_path):
    rows = [
        {"Account": "BUF", "cusip": "N/A", "ticker": "SPX-C4800", "name": "opt",
         "assetCat": "DE", "derivCat": "OPT", "lei": "", "isin": ""},
        {"Account": "CGOV", "cusip": "912797UL9", "ticker": "", "name": "UST",
         "assetCat": "DBT", "issuerCat": "UST", "derivCat": "",
         "lei": "254900HROIFWPRGM1V77", "isin": "", "invCountry": "US"},
    ]
    # Options have no Bloomberg spec; US treasuries get exactly one formula (isin)
    # keyed by "<cusip> Govt" — LEI/country/maturity/rate are already known.
    assert apply_bloomberg_formulas(rows) == 1
    assert not rows[0].get("bbgid")
    assert rows[1]["bbgid"] == "912797UL9 Govt"
    ws = _equity_master(rows, tmp_path)
    hdr = [c.value for c in ws[1]]
    idx = {c: i + 1 for i, c in enumerate(hdr)}
    # Option: no formula cells at all.
    for f in ("isin", "lei", "invCountry"):
        assert not str(ws.cell(row=2, column=idx[f]).value).startswith("=")
    # Treasury (row 3): isin is a bare BDP formula; lei stays the known literal.
    assert ws.cell(row=3, column=idx["isin"]).value == '=BDP($B3,"ID_ISIN")'
    assert ws.cell(row=3, column=idx["lei"]).value == "254900HROIFWPRGM1V77"
    # cusip is never a formula here (no custodian sheet linked).
    for excel_row in (2, 3):
        assert not str(ws.cell(row=excel_row, column=idx["cusip"]).value).startswith("=")


def test_bloomberg_error_junk_is_cleaned_in_seed(tmp_path):
    # A per-fund CSV with "#N/A N/A" junk in cusip/lei for a foreign CINS name,
    # plus a custodian that has the real CINS.
    funds = tmp_path / "funds"
    _write_per_fund_csv(funds, "av", list(EQUITY_HEADERS), [
        {"name": "AerCap Holdings NV", "lei": "#N/A N/A", "title": "AerCap Holdings NV",
         "cusip": "#N/A N/A", "isin": "NL0000687663", "ticker": "AER",
         "invCountry": "NL", "assetCat": "EC", "issuerCat": "CORP"},
    ])
    cust = _write_custodian(tmp_path, [_eq("AV", "AER", "N00985106", "AerCap Holdings NV")])

    master = tmp_path / "master.xlsx"
    seed_master_from_per_fund(funds, cust, master)
    rows, _ = read_master_xlsx(master)
    r = rows[0]
    # cusip mirrors the custodian verbatim via the =custodian! reference (the raw
    # CINS), NOT the "#N/A N/A" junk and not a formula. The N/A-for-foreign rule
    # is applied downstream by the build.
    assert r["cusip"] == "N00985106"
    # No "#N/A ..." junk survives anywhere in the workbook.
    from openpyxl import load_workbook
    ws = load_workbook(master).active
    assert not any(isinstance(c, str) and c.startswith("#")
                   for row in ws.iter_rows(values_only=True) for c in row)
    # isin/lei/invCountry are Bloomberg-owned: they become bare =BDP formulas
    # (overwriting any provided value — no fallback), and read back EMPTY until
    # resolved on a Bloomberg terminal, so failures stay visible.
    hdr = [c.value for c in ws[1]]
    for f, mnem in (("lei", "LEGAL_ENTITY_IDENTIFIER"), ("isin", "ID_ISIN"),
                    ("invCountry", "CNTRY_OF_DOMICILE")):
        cell = ws.cell(row=2, column=hdr.index(f) + 1).value
        assert cell.startswith("=BDP(") and mnem in cell and "IFERROR" not in cell
    assert r["lei"] == "" and r["isin"] == "" and r["invCountry"] == ""


def test_reference_cells_are_formulas_general_format(tmp_path):
    from openpyxl import load_workbook
    master = tmp_path / "master.xlsx"
    cust = _write_custodian(tmp_path, [_eq("FDRS", "AAPL", "037833100", "Apple Inc")])
    refresh_master(parse_custodian_csv(cust), master)

    ws = load_workbook(master).active
    hdr = [c.value for c in ws[1]]
    # isin/lei/invCountry are BDP formulas -> General so Excel evaluates them.
    for f in ("lei", "isin", "invCountry"):
        cell = ws.cell(row=2, column=hdr.index(f) + 1)
        assert str(cell.value).startswith("="), f
        assert cell.number_format != "@", f
    # cusip is a direct cross-sheet reference into the custodian sheet (no
    # XLOOKUP/LET — that's what broke loading), e.g. "=custodian!D2".
    cu = ws.cell(row=2, column=hdr.index("cusip") + 1).value
    assert cu.startswith("=custodian!"), cu
    # custodian sheet exists and read resolves cusip to the real value.
    assert "custodian" in load_workbook(master).sheetnames
    back, _ = read_master_xlsx(master)
    assert back[0]["cusip"] == "037833100"
    # ticker and name stay literal.
    for f in ("ticker", "name"):
        assert not str(ws.cell(row=2, column=hdr.index(f) + 1).value).startswith("=")


def _cash_custodian(account):
    return {"Date": "06/01/2026", "Account": account, "StockTicker": "Cash&Other",
            "CUSIP": "Cash&Other", "SecurityName": "Cash & Other", "Shares": "0",
            "Price": "1", "MarketValue": "500", "Weightings": "0.50%",
            "NetAssets": "100000", "SharesOutstanding": "0", "CreationUnits": "0",
            "MoneyMarketFlag": ""}


def test_two_sheets_and_one_to_one_with_custodian(tmp_path):
    # All custodian rows (incl. cash) appear in the master, in order, so the
    # cusip references line up; the custodian sheet holds the CSV verbatim.
    cust = _write_custodian(tmp_path, [
        _eq("FDRS", "NVDA", "67066G104", "NVIDIA Corp"),
        _cash_custodian("FDRS"),
        _eq("FDRS", "ABNB", "009066101", "Airbnb Inc"),
    ])
    master = tmp_path / "master.xlsx"
    refresh_master(parse_custodian_csv(cust), master)

    wb = load_workbook(master)
    assert wb.sheetnames == ["master", "custodian"]
    # 1:1: 3 custodian data rows -> 3 master data rows, same order.
    assert wb["master"].max_row == 4 and wb["custodian"].max_row == 4

    rows, _ = read_master_xlsx(master)
    assert len(rows) == 3
    # rawTicker mirrors the custodian StockTicker for every row (incl. cash).
    assert [r["rawTicker"] for r in rows] == ["NVDA", "Cash&Other", "ABNB"]
    # cusip mirrors the custodian column verbatim, by row index.
    assert [r["cusip"] for r in rows] == ["67066G104", "Cash&Other", "009066101"]


def test_cash_row_excluded_from_split(tmp_path):
    cust = _write_custodian(tmp_path, [
        _eq("FDRS", "NVDA", "67066G104", "NVIDIA Corp"),
        _cash_custodian("FDRS"),
    ])
    master = tmp_path / "master.xlsx"
    refresh_master(parse_custodian_csv(cust), master)
    funds = tmp_path / "funds"
    split_master(master, funds)
    header, rows = _read_csv(funds / "fdrs" / "security_master.csv")
    # Cash is inert (no assetCat/derivCat) — only the equity is written, and
    # rawTicker/bbgid/Account never leak into the per-fund CSV.
    assert len(rows) == 1 and rows[0]["ticker"] == "NVDA"
    assert "rawTicker" not in header and "bbgid" not in header and "Account" not in header


def test_no_formulas_flag_leaves_blanks(tmp_path):
    master = tmp_path / "master.xlsx"
    cust = _write_custodian(tmp_path, [_eq("FDRS", "AAPL", "037833100", "Apple Inc")])
    stats = refresh_master(parse_custodian_csv(cust), master, formulas=False)
    assert stats["formulas"] == 0
    rows, _ = read_master_xlsx(master)
    assert rows[0]["isin"] == ""  # left blank, no formula


# ── Per-asset-type Bloomberg formulas (bonds, swaps) ──────────


def _bond_custodian(account, cusip, name):
    return {"Date": "06/01/2026", "Account": account, "StockTicker": "",
            "CUSIP": cusip, "SecurityName": name, "Shares": "100000",
            "Price": "98.5", "MarketValue": "98500", "Weightings": "2.50%",
            "NetAssets": "1000000", "SharesOutstanding": "0", "CreationUnits": "0",
            "MoneyMarketFlag": ""}


def _swap_custodian(account, ticker, name):
    return {"Date": "06/01/2026", "Account": account, "StockTicker": ticker,
            "CUSIP": "N/A", "SecurityName": name, "Shares": "1", "Price": "0",
            "MarketValue": "0", "Weightings": "0.00%", "NetAssets": "1000000",
            "SharesOutstanding": "0", "CreationUnits": "0", "MoneyMarketFlag": ""}


def _formula_cell(ws, hdr, col, excel_row=2):
    return ws.cell(row=excel_row, column=hdr.index(col) + 1).value


def test_corporate_bond_formulas_keyed_corp(tmp_path):
    cust = _write_custodian(tmp_path, [
        _bond_custodian("CHYG", "00081TAK4", "ACCO Brands Corp 4.25% 03/15/2029"),
    ])
    master = tmp_path / "master.xlsx"
    refresh_master(parse_custodian_csv(cust), master)
    ws = load_workbook(master)["master"]
    hdr = [c.value for c in ws[1]]
    bb = f"${chr(ord('A') + hdr.index('bbgid'))}2"
    assert _formula_cell(ws, hdr, "bbgid") == "00081TAK4 Corp"
    # Identity + C.9 are bare BDP formulas — no IFERROR.
    assert _formula_cell(ws, hdr, "isin") == f'=BDP({bb},"ID_ISIN")'
    assert _formula_cell(ws, hdr, "lei") == f'=BDP({bb},"LEGAL_ENTITY_IDENTIFIER")'
    assert _formula_cell(ws, hdr, "invCountry") == f'=BDP({bb},"CNTRY_OF_DOMICILE")'
    assert _formula_cell(ws, hdr, "maturityDt") == f'=TEXT(BDP({bb},"MATURITY"),"yyyy-mm-dd")'
    assert _formula_cell(ws, hdr, "annualizedRt") == f'=BDP({bb},"CPN")'
    assert _formula_cell(ws, hdr, "couponKind") == f'=BDP({bb},"CPN_TYP")'
    for f in ("isin", "lei", "maturityDt", "couponKind"):
        assert "IFERROR" not in _formula_cell(ws, hdr, f)
    # assetCat/issuerCat are the locally-known literals.
    assert _formula_cell(ws, hdr, "assetCat") == "DBT"
    assert _formula_cell(ws, hdr, "issuerCat") == "CORP"
    # Unresolved off-terminal → C.9/identity read back empty (visible failure).
    rows, _ = read_master_xlsx(master)
    assert rows[0]["maturityDt"] == "" and rows[0]["lei"] == "" and rows[0]["invCountry"] == ""


def test_swap_reference_identity_formulas(tmp_path):
    cust = _write_custodian(tmp_path, [
        _swap_custodian("CMAG", "02079K305-TRS-05/31/27-L-CANT",
                        "ALPHABET INC.-SWAP-CANT-L"),
    ])
    master = tmp_path / "master.xlsx"
    refresh_master(parse_custodian_csv(cust), master)
    ws = load_workbook(master)["master"]
    hdr = [c.value for c in ws[1]]
    bb = f"${chr(ord('A') + hdr.index('bbgid'))}2"
    # Keyed by the reference security's CUSIP (parsed from the swap ticker).
    assert _formula_cell(ws, hdr, "bbgid") == "02079K305 Equity"
    assert _formula_cell(ws, hdr, "refIsin") == f'=BDP({bb},"ID_ISIN")'
    assert _formula_cell(ws, hdr, "refTicker") == f'=BDP({bb},"TICKER")'
    assert _formula_cell(ws, hdr, "refIssuerName") == f'=BDP({bb},"ISSUER")'
    assert _formula_cell(ws, hdr, "refIssueTitle") == f'=BDP({bb},"NAME")'
    # Economics stay operator-entered, NOT formulas.
    for f in ("counterpartyLei", "notionalAmt", "unrealizedAppr"):
        assert not str(_formula_cell(ws, hdr, f)).startswith("=")


def test_coupon_kind_normalization():
    assert _normalize_coupon_kind("FIXED") == "Fixed"
    assert _normalize_coupon_kind("FLOATING") == "Floating"
    assert _normalize_coupon_kind("ZERO COUPON") == "None"
    assert _normalize_coupon_kind("") == ""
    # Unmapped value passes through unchanged so validation flags it.
    assert _normalize_coupon_kind("STEP CPN") == "STEP CPN"


def test_spec_selection():
    assert _bbg_spec_key({"assetCat": "EC"}) == "EC"
    assert _bbg_spec_key({"assetCat": "STIV"}) == "STIV"
    assert _bbg_spec_key({"assetCat": "DBT", "issuerCat": "UST"}) == "DBT_UST"
    assert _bbg_spec_key({"assetCat": "DBT", "issuerCat": "CORP"}) == "DBT_CORP"
    assert _bbg_spec_key({"assetCat": "DE", "derivCat": "SWP", "refCusip": "123"}) == "SWP_REF"
    assert _bbg_spec_key({"assetCat": "DE", "derivCat": "OPT"}) is None
