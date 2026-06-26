"""One global master spreadsheet → per-fund security masters.

Instead of editing ~90 per-fund ``data/funds/<fund>/security_master.csv`` files,
the operator maintains a single Excel workbook
(``data/master/security_master.xlsx``) with two sheets:

  * ``custodian`` — the US Bank custodian CSV in xlsx, verbatim. Never edited.
  * ``master``    — one row per custodian row (strictly 1:1, same order),
                    carrying the union of every enrichment column plus live
                    Bloomberg ``=BDP(...)`` formulas (equities, for now).

Because the sheets are row-aligned, each ``cusip`` cell is a direct cross-sheet
reference ``=custodian!<col><row>`` — copied from the custodian, never looked up.
The build pipeline is untouched — a ``split`` step regenerates the per-fund CSVs
from the master.

Why .xlsx: identifier columns (CUSIP/ISIN/...) are stored with Excel Text
formatting (``number_format='@'``), which structurally prevents the
scientific-notation / dropped-leading-zero CUSIP corruption that plain CSVs
suffer when opened in Excel.

Three operations:
  * ``refresh_master``  — add new holdings, drop gone ones, KEEP manual edits.
  * ``split_master``    — project each fund's rows back to per-fund CSVs.
  * ``seed_master_from_per_fund`` — one-time migration from the existing CSVs.

Everything reuses the per-type entry builders and keying in ``custodian.py`` so
there is exactly one source of truth for classification and XML-default values.
"""

import csv
import logging
import os
import re
import tempfile
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from nport.config import _HOLDINGS_KEY_MAP
from nport.custodian import (
    EQUITY_HEADERS,
    OPTION_HEADERS,
    SWAP_HEADERS,
    HoldingType,
    build_corporate_bond_entry,
    build_equity_entry,
    build_mm_entry,
    build_option_entry,
    build_swap_entry,
    build_treasury_entry,
    classify_holding,
    load_xml_reference,
    write_security_master,
    _sm_entry_key,
    _sm_lookup_key,
)
from nport.cusip import normalize_cusip

logger = logging.getLogger(__name__)


# ── Schema ────────────────────────────────────────────────────

# Master-only columns, shown first, and dropped on split (per-fund CSVs begin at
# ``name``). ``Account`` groups the split; ``bbgid`` is the Bloomberg security
# string ("<ticker> US Equity") the equity formulas reference; ``rawTicker`` is
# the verbatim custodian StockTicker, kept for audit (the processed ``ticker``
# can differ, e.g. options carry a synthetic id).
IDENTITY_COLUMNS = ["Account", "bbgid", "rawTicker"]

# Sheet 1 is the custodian CSV verbatim. Its column order (US Bank's) — used to
# locate the CUSIP column the master's ``=custodian!…`` references point at.
_CUSTODIAN_SHEET = "custodian"
CUSTODIAN_HEADER = [
    "Date", "Account", "StockTicker", "CUSIP", "SecurityName", "Shares", "Price",
    "MarketValue", "Weightings", "NetAssets", "SharesOutstanding", "CreationUnits",
    "MoneyMarketFlag",
]


def _custodian_values(cr) -> list[str]:
    """A CustodianRow as a list in ``CUSTODIAN_HEADER`` order."""
    return [
        cr.date, cr.account, cr.stock_ticker, cr.cusip, cr.security_name, cr.shares,
        cr.price, cr.market_value, cr.weightings, cr.net_assets,
        cr.shares_outstanding, cr.creation_units, cr.money_market_flag,
    ]

# Debt columns (the bond_fund shape adds the first three; treasuries normally
# stay in the equity shape).
DEBT_COLUMNS = [
    "maturityDt", "couponKind", "annualizedRt",
    "isDefault", "areIntrstPmntsInArrs", "isPaidKind",
]


def _build_enrichment_columns() -> list[str]:
    """Canonical union of every per-fund security_master column.

    Order: the 9 equity base cols, then the option block, then the swap block,
    then debt cols. Built from the header constants (which already match what
    the entry builders emit) so it cannot drift from ``custodian.py``.
    """
    cols: list[str] = []
    for seq in (EQUITY_HEADERS, OPTION_HEADERS, SWAP_HEADERS, DEBT_COLUMNS):
        for c in seq:
            if c not in cols:
                cols.append(c)
    # Every column must be a legal holdings field.
    assert all(c in _HOLDINGS_KEY_MAP for c in cols), \
        [c for c in cols if c not in _HOLDINGS_KEY_MAP]
    return cols


MASTER_ENRICHMENT_COLUMNS = _build_enrichment_columns()

# Columns stored as Excel Text to defeat spreadsheet corruption.
ID_TEXT_COLUMNS = {
    "cusip", "isin", "ticker", "rawTicker", "lei", "counterpartyLei",
    "refCusip", "refIsin", "refTicker",
}

# ── Bloomberg enrichment, per asset type ──────────────────────
#
# Each asset type pulls the reference fields Bloomberg can supply, via a single
# security id written into the row's ``bbgid`` cell. A spec is
# ``(key_fn, fields)`` where ``key_fn(row_dict)`` builds the Bloomberg security
# string and ``fields`` maps a master column → ``(BDP mnemonic, kind)``.
#
# Formulas are BARE ``=BDP(...)`` with NO fallback: a failed lookup shows the
# live ``#N/A`` so the operator sees and fixes it. The master is populated once
# on a Bloomberg terminal and then exported; a hidden fallback would let bad data
# ship silently.
#
# CUSIP is never here: it is a live ``=custodian!…`` reference into sheet 1.
# name/title are excluded too (the custodian's are clean; Bloomberg SECURITY_NAME
# is ALL-CAPS / cryptic). Mnemonics were verified live on the terminal
# (LEGAL_ENTITY_IDENTIFIER — NOT ID_LEI; CNTRY_OF_DOMICILE — issuer home country,
# matching the real filings: SPOT→SE, NU→BR, TEAM→AU, RPRX→US).
#
# kind: "value" passthrough | "date" (BDP date serial → yyyy-mm-dd) |
#       "couponKind" (CPN_TYP normalized to the N-PORT enum on read).
BLOOMBERG_SPECS = {
    "EC": (
        lambda r: f"{(r.get('ticker') or '').strip()} US Equity",
        {"isin": ("ID_ISIN", "value"),
         "lei": ("LEGAL_ENTITY_IDENTIFIER", "value"),
         "invCountry": ("CNTRY_OF_DOMICILE", "value")},
    ),
    "STIV": (
        lambda r: f"{(r.get('ticker') or '').strip()} US Equity",
        {"isin": ("ID_ISIN", "value"),
         "lei": ("LEGAL_ENTITY_IDENTIFIER", "value"),
         "invCountry": ("CNTRY_OF_DOMICILE", "value")},
    ),
    # US Treasuries (notes, bonds AND bills) pull the full identity + C.9 set
    # under the Govt key — verified live: MATURITY/CPN/CPN_TYP resolve for notes
    # (CPN_TYP=FIXED) and bills (CPN_TYP=ZERO, CPN=0); LEI is the US Treasury LEI,
    # CNTRY_OF_DOMICILE=US.
    "DBT_UST": (
        lambda r: f"{(r.get('cusip') or '').strip()} Govt",
        {"isin": ("ID_ISIN", "value"),
         "lei": ("LEGAL_ENTITY_IDENTIFIER", "value"),
         "invCountry": ("CNTRY_OF_DOMICILE", "value"),
         "maturityDt": ("MATURITY", "date"),
         "annualizedRt": ("CPN", "value"),
         "couponKind": ("CPN_TYP", "couponKind")},
    ),
    "DBT_CORP": (
        lambda r: f"{(r.get('cusip') or '').strip()} Corp",
        {"isin": ("ID_ISIN", "value"),
         "lei": ("LEGAL_ENTITY_IDENTIFIER", "value"),
         "invCountry": ("CNTRY_OF_DOMICILE", "value"),
         "maturityDt": ("MATURITY", "date"),
         "annualizedRt": ("CPN", "value"),
         "couponKind": ("CPN_TYP", "couponKind")},
    ),
    # Swaps: key off the REFERENCE security (its CUSIP, parsed from the swap
    # ticker) and fill only the reference-instrument identity. Contract economics
    # (counterparty LEI, leg rates/spread, notional, unrealized) are not on
    # Bloomberg and stay operator-entered.
    "SWP_REF": (
        lambda r: f"{(r.get('refCusip') or '').strip()} Equity",
        {"refIsin": ("ID_ISIN", "value"),
         "refTicker": ("TICKER", "value"),
         "refIssuerName": ("ISSUER", "value"),
         "refIssueTitle": ("NAME", "value")},
    ),
}

# Union of every column any spec turns into a formula — used by the read/write
# loops to know which cells are Bloomberg-owned.
_BBG_FORMULA_COLUMNS = {col for _kf, fields in BLOOMBERG_SPECS.values() for col in fields}

_NAME_MAX_LEN = 30  # XSD <name> max length

# Schema-valid default for a Bloomberg-owned field that comes back empty/#N/A.
# Only fields with a meaningful N-PORT sentinel are listed: an issuer with no LEI
# in Bloomberg is reported "N/A" (verified: no Bloomberg/BQL route to it here),
# and a missing domicile defaults to US. isin/maturityDt/couponKind/annualizedRt
# have NO default — a blank there is a real gap and must fail validation visibly.
_FIELD_DEFAULT = {"lei": "N/A", "invCountry": "US"}

# Date columns normalized to ISO YYYY-MM-DD on read. maturityDt comes from
# Bloomberg via Excel `=TEXT(BDP(...))`, which can leak a locale string like
# "9/3/2026" when BDP returns text; expDt/terminationDt are custodian-parsed.
_DATE_COLUMNS = {"maturityDt", "expDt", "terminationDt"}

# Bloomberg CPN_TYP → N-PORT couponKind enum (Fixed/Floating/Variable/None). An
# unmapped non-empty value passes through unchanged so validation flags it. A
# pay-in-kind bond is fixed-rate paid in kind → Fixed (with isPaidKind=Y).
_COUPON_KIND_MAP = {
    "FIXED": "Fixed", "FLOATING": "Floating", "VARIABLE": "Variable",
    "NONE": "None", "ZERO": "None", "ZERO COUPON": "None", "ZERO CPN": "None",
    "PAY-IN-KIND": "Fixed", "PAY IN KIND": "Fixed", "PIK": "Fixed",
}


def _normalize_date(value: str) -> str:
    """Coerce a date string to ISO YYYY-MM-DD.

    Bloomberg's Excel add-in can return a date as locale text ("9/3/2026") that
    `TEXT(...,"yyyy-mm-dd")` fails to reformat; a cached date serial round-trips
    as "2026-09-03 00:00:00". Both (and already-ISO values) are normalized; an
    unparseable value is returned unchanged so it stays visible.
    """
    s = (value or "").strip()
    if not s:
        return ""
    if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
        return s
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d", "%m-%d-%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s


def _normalize_coupon_kind(value: str) -> str:
    v = (value or "").strip()
    if not v:
        return ""
    return _COUPON_KIND_MAP.get(v.upper(), v)


def _bbg_spec_key(row: dict[str, str]) -> str | None:
    """Which ``BLOOMBERG_SPECS`` entry applies to this master row (or None)."""
    if (row.get("derivCat") or "").strip() == "SWP" and (row.get("refCusip") or "").strip():
        return "SWP_REF"
    asset = (row.get("assetCat") or "").strip()
    if asset == "EC":
        return "EC"
    if asset == "STIV":
        return "STIV"
    if asset == "DBT":
        return "DBT_UST" if (row.get("issuerCat") or "").strip() == "UST" else "DBT_CORP"
    return None  # options and anything else: no Bloomberg lookup


def _bloomberg_formula(mnemonic: str, key_cell: str, kind: str) -> str:
    """A BARE Bloomberg formula referencing the row's key cell — no fallback.

    ``value``/``couponKind`` → ``=BDP($B2,"MN")``; ``date`` →
    ``=TEXT(BDP($B2,"MN"),"yyyy-mm-dd")``. A failed/empty lookup shows the live
    Bloomberg error in the cell, by design — see ``BLOOMBERG_SPECS``.
    """
    call = f'BDP({key_cell},"{mnemonic}")'
    if kind == "date":
        return f'=TEXT({call},"yyyy-mm-dd")'
    return f"={call}"


def _is_formula(value: str) -> bool:
    return isinstance(value, str) and value.startswith("=")


def _strip_bbg_error(value: str) -> str:
    """Blank out Bloomberg-error text saved as a literal by a prior session.

    Spreadsheet round-trips leave strings like ``#N/A N/A``,
    ``#N/A Invalid Security`` or ``#N/A Field Not Applicable`` in cells. They
    are not data — treat them as missing so keying, the custodian match, and the
    gap-fill all behave as if the cell were empty.
    """
    if isinstance(value, str) and value.startswith("#"):
        return ""
    return value


def apply_bloomberg_formulas(rows: list[dict[str, str]], overwrite: bool = True) -> int:
    """Set the Bloomberg lookup key (``bbgid``) for every row that has a spec.

    Per ``BLOOMBERG_SPECS``: equities/money-market key by ``"<ticker> US
    Equity"``, bonds by ``"<cusip> Corp"``/``"<cusip> Govt"``, and swaps by their
    reference security ``"<refCusip> Equity"``. ``write_master_xlsx`` then turns
    each spec column into a bare ``=BDP(<bbgid>, ...)`` formula. Options (and any
    other type) have no spec and are left untouched.

    No fallback values are computed or stored — a failed Bloomberg lookup must
    stay visible (see ``BLOOMBERG_SPECS``). ``overwrite`` is accepted for call
    compatibility; spec cells are always formula-driven.

    Returns the number of formula cells that will be written.
    """
    count = 0
    for row in rows:
        spec_key = _bbg_spec_key(row)
        if spec_key is None:
            continue
        key_fn, fields = BLOOMBERG_SPECS[spec_key]
        bbg = key_fn(row)
        ident = bbg.split(" ", 1)[0]  # the ticker/cusip/refCusip token
        if not ident or ident == "N/A" or ident.startswith("#"):
            continue
        row["bbgid"] = bbg.strip()
        count += len(fields)
    return count


def _ordered_enrichment(extra_cols: set[str]) -> list[str]:
    """Enrichment columns in canonical order, with any legal extras appended.

    Hand-divergent funds (e.g. ``leveraged_etf``'s granular receive-leg
    columns) carry legal columns outside the canonical set; preserve them in
    ``_HOLDINGS_KEY_MAP`` order so nothing is lost.
    """
    cols = list(MASTER_ENRICHMENT_COLUMNS)
    for c in _HOLDINGS_KEY_MAP:
        if c in extra_cols and c not in cols:
            cols.append(c)
    return cols


def _finalize_header(rows: list[dict]) -> list[str]:
    """Master header covering every legal enrichment column present in rows."""
    extra = {
        c for r in rows for c in r
        if c in _HOLDINGS_KEY_MAP and c not in MASTER_ENRICHMENT_COLUMNS
    }
    return IDENTITY_COLUMNS + _ordered_enrichment(extra)


# ── Row construction ──────────────────────────────────────────


def _build_entry(ht: HoldingType, row, ref: dict) -> dict[str, str] | None:
    """Build the narrow per-type security-master entry for a custodian row."""
    if ht == HoldingType.EQUITY:
        return build_equity_entry(row.stock_ticker, row.security_name, row.cusip, ref)
    if ht == HoldingType.MONEY_MARKET:
        return build_mm_entry(row)
    if ht == HoldingType.OPTION:
        return build_option_entry(row)
    if ht == HoldingType.SWAP:
        return build_swap_entry(row)
    if ht == HoldingType.TREASURY:
        return build_treasury_entry(row)
    if ht == HoldingType.CORPORATE_BOND:
        return build_corporate_bond_entry(row)
    return None


# ── xlsx I/O ──────────────────────────────────────────────────


def _cell_to_str(value) -> str:
    """Coerce an openpyxl cell value to a string without float artifacts."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _read_custodian_cusips(wb) -> list[str]:
    """The CUSIP column of sheet 1, in row order (1:1 with the master rows)."""
    if _CUSTODIAN_SHEET not in wb.sheetnames:
        return []
    crows = list(wb[_CUSTODIAN_SHEET].iter_rows(values_only=True))
    if not crows:
        return []
    chdr = [_cell_to_str(h).strip() for h in crows[0]]
    if "CUSIP" not in chdr:
        return []
    ci = chdr.index("CUSIP")
    return [_cell_to_str(cr[ci]) if cr and ci < len(cr) else "" for cr in crows[1:]]


def read_master_xlsx(path: Path) -> tuple[list[dict[str, str]], list[str]]:
    """Read the master workbook into row dicts (all values as strings).

    Reads cached formula results (``data_only``) and, in parallel, the raw
    formulas. A Bloomberg cell whose formula has no cached value yet (the file
    was never opened on a Bloomberg machine) resolves to that field's fallback
    so downstream output stays schema-valid (e.g. ``lei`` → ``N/A``).

    CUSIP is a ``=custodian!…`` reference openpyxl can't evaluate, so it is
    resolved from sheet 1 by row index — the master is built strictly 1:1 with
    the custodian, so master data row *i* maps to custodian data row *i*.
    """
    wb_v = load_workbook(path, data_only=True)
    wb_f = load_workbook(path, data_only=False)

    def _master(wb):
        return wb["master"] if "master" in wb.sheetnames else wb.active

    rows_v = list(_master(wb_v).iter_rows(values_only=True))
    if not rows_v:
        return [], []
    header = [(_cell_to_str(h)).strip() for h in rows_v[0] if h is not None]

    cust_cusips = _read_custodian_cusips(wb_f)  # by row index, 1:1 with master rows
    data_v = rows_v[1:]

    rows: list[dict[str, str]] = []
    for idx, vraw in enumerate(data_v):
        if vraw is None or all(c is None for c in vraw):
            continue
        rec: dict[str, str] = {}
        for i, col in enumerate(header):
            sval = _cell_to_str(vraw[i]) if i < len(vraw) else ""
            if col in _BBG_FORMULA_COLUMNS:
                # Bloomberg-owned cell. Drop the literal "#N/A …" junk a failed
                # lookup leaves, map couponKind/normalize dates, then apply the
                # schema-valid default where one exists (lei→N/A, invCountry→US).
                # Fields without a default stay blank so a real gap (e.g. a bond
                # with no maturity) still fails validation visibly.
                sval = _strip_bbg_error(sval)
                if col == "couponKind":
                    sval = _normalize_coupon_kind(sval)
                elif col in _DATE_COLUMNS:
                    sval = _normalize_date(sval)
                if not sval and col in _FIELD_DEFAULT:
                    sval = _FIELD_DEFAULT[col]
                rec[col] = sval
                continue
            sval = _strip_bbg_error(sval)  # drop literal "#N/A ..." junk from manual cells
            if col in _DATE_COLUMNS:
                sval = _normalize_date(sval)
            rec[col] = sval
        # CUSIP: the live link into sheet 1, resolved here by row index (since
        # openpyxl can't evaluate the formula). Falls back to defensive repair
        # for a standalone master without the custodian sheet (unit tests).
        if cust_cusips and idx < len(cust_cusips):
            rec["cusip"] = cust_cusips[idx]
        else:
            rec["cusip"], warn = normalize_cusip(rec.get("cusip", ""), rec.get("isin", ""))
            if warn:
                logger.warning("%s: %s", rec.get("name", rec.get("ticker", "?")), warn)
        if "refCusip" in rec:
            rec["refCusip"], _ = normalize_cusip(rec.get("refCusip", ""), rec.get("refIsin", ""))
        rows.append(rec)
    return rows, header


def write_master_xlsx(
    rows: list[dict[str, str]],
    header: list[str],
    path: Path,
    custodian_rows: list | None = None,
) -> None:
    """Write a two-sheet master workbook.

    Sheet ``master`` holds ``rows`` (the enrichment). Sheet ``custodian`` holds
    the raw custodian CSV (``custodian_rows``) verbatim. The two sheets are
    strictly 1:1 by row, so each ``cusip`` cell is a direct cross-sheet
    reference ``=custodian!<CUSIPcol><row>`` — copied from the custodian, never
    looked up. Equity rows (``assetCat == "EC"`` with a ``bbgid``) get
    ``=BDP(<bbgid>, ...)`` formulas for isin/lei/invCountry. Everything else is
    written literally. Without ``custodian_rows`` (unit tests), cusip is written
    as the row's literal value and no custodian sheet is added.
    """
    from openpyxl.utils import get_column_letter

    link = custodian_rows is not None
    if link and len(custodian_rows) != len(rows):
        raise ValueError(
            f"master/custodian row mismatch: {len(rows)} vs {len(custodian_rows)} "
            "(the master must be built 1:1 with the custodian)"
        )
    cust_cusip_col = (
        get_column_letter(CUSTODIAN_HEADER.index("CUSIP") + 1) if link else None
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "master"
    ws.append(header)

    bbgid_pos = header.index("bbgid") + 1 if "bbgid" in header else None

    for idx, r in enumerate(rows):
        excel_row = idx + 2
        spec_key = _bbg_spec_key(r)
        fields = BLOOMBERG_SPECS[spec_key][1] if spec_key else {}
        bbgid_ref = (f"${get_column_letter(bbgid_pos)}{excel_row}"
                     if (bbgid_pos and fields and (r.get('bbgid') or '').strip()) else None)
        out = []
        for col in header:
            val = r.get(col, "")
            if col == "cusip" and link:
                val = f"={_CUSTODIAN_SHEET}!{cust_cusip_col}{excel_row}"
            elif bbgid_ref and col in fields:
                mnemonic, kind = fields[col]
                val = _bloomberg_formula(mnemonic, bbgid_ref, kind)
            out.append(val)
        ws.append(out)

    # Stamp Text format on identifier cells so Excel keeps them as text — but
    # NOT on formula cells, which must stay General so Excel evaluates them.
    text_cols = [i + 1 for i, c in enumerate(header) if c in ID_TEXT_COLUMNS]
    for col_idx in text_cols:
        for excel_row in range(2, ws.max_row + 1):
            cell = ws.cell(row=excel_row, column=col_idx)
            if not _is_formula(cell.value):
                cell.number_format = "@"

    # Sheet 1: the custodian CSV verbatim (the cusip references point here).
    if link:
        cs = wb.create_sheet(_CUSTODIAN_SHEET)
        cs.append(CUSTODIAN_HEADER)
        for cr in custodian_rows:
            cs.append(_custodian_values(cr))
        tkr_c = CUSTODIAN_HEADER.index("StockTicker") + 1
        cus_c = CUSTODIAN_HEADER.index("CUSIP") + 1
        for excel_row in range(2, cs.max_row + 1):
            cs.cell(row=excel_row, column=tkr_c).number_format = "@"
            cs.cell(row=excel_row, column=cus_c).number_format = "@"

    path.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp = tempfile.mkstemp(dir=path.parent, suffix=".xlsx")
    os.close(fd)
    try:
        wb.save(tmp)
        Path(tmp).replace(path)
    except Exception:
        Path(tmp).unlink(missing_ok=True)
        raise


# ── Row assembly (custodian → master, 1:1) ────────────────────


def _assemble_master_rows(
    custodian_rows: list,
    manual_by_key: dict[tuple[str, str], dict[str, str]],
    ref: dict | None = None,
    formulas: bool = True,
    overwrite_formulas: bool = False,
) -> tuple[list[dict[str, str]], int]:
    """Build master rows strictly 1:1 with the custodian (same order, all rows).

    Every custodian row produces exactly one master row, so the two sheets stay
    row-aligned (the ``=custodian!…`` cusip references depend on this). For each
    non-cash holding the enrichment is taken from ``manual_by_key`` (operator-
    entered values from the existing master or the per-fund CSVs) keyed by
    ``(Account, security-key)``, falling back to a fresh build from the
    custodian. Cash rows carry blank enrichment (the build skips them). Equity
    rows then get live ``=BDP(...)`` formulas.

    Returns ``(rows, n_formula_cells)``.
    """
    ref = ref or {}
    result: list[dict[str, str]] = []
    for row in custodian_rows:
        ht = classify_holding(row)
        account = row.account.upper()
        wide = {c: "" for c in MASTER_ENRICHMENT_COLUMNS}
        if ht != HoldingType.CASH:
            key = _sm_lookup_key(ht, row)
            existing = manual_by_key.get((account, key)) if key else None
            if ht in (HoldingType.OPTION, HoldingType.SWAP):
                # Derivative economics (counterparty/LEI/notional/legs) are deterministic
                # from the custodian — always rebuild so code fixes propagate; carry over
                # only the truly-manual fields the operator enters (delta, unrealizedAppr).
                entry = _build_entry(ht, row, ref)
                for mf in ("delta", "unrealizedAppr"):
                    if existing and (existing.get(mf) or "").strip():
                        entry[mf] = existing[mf]
            else:
                entry = existing if existing is not None else _build_entry(ht, row, ref)
            if entry:
                wide.update(entry)
            # Equity ticker is the custodian StockTicker (source of truth) — never
            # a stale per-fund value. Options/swaps keep their synthetic-id ticker.
            if ht == HoldingType.EQUITY:
                wide["ticker"] = (row.stock_ticker or "").strip()
        wide["Account"] = account
        wide["rawTicker"] = (row.stock_ticker or "").strip()
        wide["cusip"] = (row.cusip or "")  # literal fallback; write links to sheet 1
        result.append(wide)

    n_formulas = (
        apply_bloomberg_formulas(result, overwrite=overwrite_formulas) if formulas else 0
    )
    return result, n_formulas


# ── Refresh (custodian → master) ──────────────────────────────


def refresh_master(
    custodian_rows: list,
    master_path: Path,
    xml_dir: Path | None = None,
    accounts: list[str] | None = None,
    formulas: bool = True,
    overwrite_formulas: bool = False,
) -> dict[str, int]:
    """Rebuild the master workbook 1:1 from the current custodian.

    The custodian is the source of truth: the master mirrors it row-for-row. We
    preserve operator-entered fields by merging the existing master's values
    (keyed by Account + security) into the matching rows; holdings new to the
    custodian are built fresh, and holdings gone from the custodian disappear.

    When ``formulas`` is set, blank Bloomberg-derivable cells are filled with
    live ``=BDP(...)`` formulas (``overwrite_formulas`` re-formulas even
    populated cells).

    Returns stats ``{"added", "kept", "removed", "formulas"}``.
    """
    ref: dict[str, dict[str, str]] = {}
    if xml_dir and Path(xml_dir).is_dir():
        ref = load_xml_reference(Path(xml_dir))

    existing_rows: list[dict[str, str]] = []
    if master_path.is_file():
        existing_rows, _ = read_master_xlsx(master_path)
    manual_by_key = {(r.get("Account", ""), _sm_entry_key(r)): r for r in existing_rows}

    target_set = (
        {a.upper() for a in accounts} if accounts
        else {r.account.upper() for r in custodian_rows}
    )

    result, n_formulas = _assemble_master_rows(
        custodian_rows, manual_by_key, ref, formulas, overwrite_formulas
    )

    # Stats over the targeted accounts: which custodian keys are new vs already
    # in the master, and which master keys are no longer in the custodian.
    cust_keys: set[tuple[str, str]] = set()
    for row in custodian_rows:
        ht = classify_holding(row)
        if ht == HoldingType.CASH:
            continue
        key = _sm_lookup_key(ht, row)
        if key:
            cust_keys.add((row.account.upper(), key))
    existing_keys = set(manual_by_key)
    stats = {
        "added": sum(1 for k in cust_keys if k not in existing_keys and k[0] in target_set),
        "kept": sum(1 for k in cust_keys if k in existing_keys and k[0] in target_set),
        "removed": sum(1 for k in existing_keys if k not in cust_keys and k[0] in target_set),
        "formulas": n_formulas,
    }
    write_master_xlsx(result, _finalize_header(result), master_path, custodian_rows)
    return stats


# ── Split (master → per-fund CSVs) ────────────────────────────


def split_master(
    master_path: Path,
    funds_dir: Path,
    accounts: list[str] | None = None,
    dry_run: bool = False,
) -> list[tuple[str, Path, int]]:
    """Write each fund's per-fund security_master.csv from the master.

    Every per-fund file is a literal projection of the master: the SAME full
    column set (``Account``/``bbgid``/``rawTicker`` + every enrichment column,
    in master order), just filtered to that fund's rows by ``Account``. Type-
    irrelevant cells stay blank, exactly as in the master. The build's loader
    maps known headers and ignores the rest, so the master-only columns are inert.

    Returns ``[(account, path, n_rows)]``.
    """
    rows, header = read_master_xlsx(master_path)
    grouped: dict[str, list[dict[str, str]]] = {}
    for r in rows:
        # Skip cash rows (no security type) — they carry no enrichment and the
        # build skips them anyway. They exist in the master only to keep it 1:1
        # with the custodian sheet.
        if not (r.get("assetCat") or r.get("derivCat")):
            continue
        grouped.setdefault(r.get("Account", ""), []).append(r)

    target = [a.upper() for a in accounts] if accounts else sorted(grouped)
    results: list[tuple[str, Path, int]] = []
    for account in target:
        arows = grouped.get(account, [])
        if not arows:
            continue
        out_rows = [{c: r.get(c, "") for c in header} for r in arows]
        path = funds_dir / account.lower() / "security_master.csv"
        if not dry_run:
            write_security_master(out_rows, header, path)
        results.append((account, path, len(out_rows)))
    return results


# ── Seed (per-fund CSVs → master) ─────────────────────────────


def seed_master_from_per_fund(
    funds_dir: Path,
    custodian_path: Path | None,
    master_path: Path,
    formulas: bool = True,
) -> dict[str, int]:
    """Build the master workbook 1:1 with the custodian, adding Bloomberg fields.

    The master mirrors the custodian CSV row-for-row (sheet 1 is that CSV, sheet
    2 the enrichment). For each holding we take its enrichment from the matching
    per-fund ``security_master.csv`` (preserving operator-entered values like
    swap counterparties); a holding new to the per-fund files is built fresh
    from the custodian. The custodian is the source of truth: funds/securities
    not in it never appear. When ``formulas`` is set, equity rows get live
    ``=BDP(...)`` formulas.
    """
    if not (custodian_path and Path(custodian_path).is_file()):
        raise FileNotFoundError(f"Custodian CSV required: {custodian_path}")
    from nport.custodian import parse_custodian_csv
    custodian_rows = parse_custodian_csv(Path(custodian_path))
    cust_accounts = {r.account.upper() for r in custodian_rows}

    # Index per-fund CSV enrichment by (Account, security-key).
    manual_by_key: dict[tuple[str, str], dict[str, str]] = {}
    csv_accounts: set[str] = set()
    for sm in sorted(funds_dir.glob("*/security_master.csv")):
        account = sm.parent.name.upper()
        csv_accounts.add(account)
        with open(sm, newline="", encoding="utf-8") as f:
            for raw in csv.DictReader(f):
                # Strip header whitespace ("issuerCat ") and blank "#N/A" junk.
                rec = {
                    (k.strip() if k else k): _strip_bbg_error(v if v is not None else "")
                    for k, v in raw.items()
                }
                entry = {k: v for k, v in rec.items() if k in _HOLDINGS_KEY_MAP}
                manual_by_key[(account, _sm_entry_key(entry))] = entry

    result, n_formulas = _assemble_master_rows(custodian_rows, manual_by_key, formulas=formulas)
    write_master_xlsx(result, _finalize_header(result), master_path, custodian_rows)

    skipped = sorted(csv_accounts - cust_accounts)
    if skipped:
        logger.warning("Skipped %d fund(s) not in custodian: %s",
                       len(skipped), ", ".join(skipped))
    non_cash = sum(1 for r in result if r.get("assetCat") or r.get("derivCat"))
    return {"funds": len(cust_accounts), "rows": len(result), "holdings": non_cash,
            "formulas": n_formulas, "skipped": skipped}
