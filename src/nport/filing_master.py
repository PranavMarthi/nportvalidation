"""Per-period filing-data master workbook → per-fund filing_data.txt.

One row per fund. The Bloomberg-derivable fields (`rtn1-3`, the 3 reporting-period
monthly total returns) are live ``=BDP(...)`` formulas; net/total assets come from the
custodian; submission flags, period dates, and balance-sheet items are constants. The
fund-accounting / transfer-agent fields (realized & unrealized gains, capital flows) have
no data feed and are written as ``0`` for the operator to fill in the sheet.

Workflow: ``build-filing-master`` writes the workbook → open it on a Bloomberg terminal so
the return formulas calculate → save → ``split-filing-master`` writes each fund's
``filings/<period>/filing_data.txt``. Mirrors the security-master master/split pattern.
"""
import calendar
import json
import os
import tempfile
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from nport.ap_orders import flows_from_csv
from nport.config import _FILING_KEY_MAP
from nport.custodian import (
    HoldingType,
    _period_end_date,
    classify_holding,
    parse_custodian_csv,
)
from nport.master_sheet import _cell_to_str

# ── Schema ────────────────────────────────────────────────────

# camelCase filing-data keys, in canonical (template) order. Risk-metric JSON blobs are
# optional and excluded — they don't belong in a flat per-fund sheet.
_RISK_KEYS = {"curMetricsJson", "creditSprdRiskIgJson", "creditSprdRiskNonigJson"}
FILING_COLUMNS = [k for k in _FILING_KEY_MAP if k not in _RISK_KEYS]

# Master-only columns (dropped on split). bbgid is the Bloomberg security the return
# formulas reference ("<ticker> US Equity").
IDENTITY_COLUMNS = ["Account", "bbgid"]
HEADER = IDENTITY_COLUMNS + FILING_COLUMNS

RETURN_COLS = ("rtn1", "rtn2", "rtn3")  # Bloomberg =BDP formulas (one month each)

# ── B.3 risk sheet (one row per debt holding; Bloomberg durations via =BDP) ──
RISK_HEADER = ["Account", "cusip", "bbgid", "valUSD", "durAdj", "spreadDur", "maturity", "ratingSP"]
# risk-sheet column -> Bloomberg mnemonic (formula cells, referencing the bbgid column)
_RISK_BDP_FIELDS = [
    ("durAdj", "DUR_ADJ_MID"), ("spreadDur", "OAS_SPREAD_DUR_MID"),
    ("maturity", "MATURITY"), ("ratingSP", "RTG_SP"),
]
# XSD maturity buckets and the JSON keys the builder consumes (see builder._RISK_PERIOD_KEYS).
_BUCKETS = ["3month", "1year", "5year", "10year", "30year"]
# Nearest-tenor cut points (geometric midpoints of 0.25/1/5/10/30 yr).
_BUCKET_CUTS = [(0.5, "3month"), (2.2360679, "1year"), (7.0710678, "5year"), (17.320508, "10year")]
# S&P investment-grade rating set (AAA … BBB-); blank/unrated treated as IG (treasuries).
_IG_RATINGS = {"AAA", "AA+", "AA", "AA-", "A+", "A", "A-", "BBB+", "BBB", "BBB-"}

# Constant (non-period) values.
_CONST = {
    "submissionType": "NPORT-P", "liveTestFlag": "TEST", "isFinalFiling": "N",
    "isNonCashCollateral": "N", "nameDesignatedIndex": "N/A", "indexIdentifier": "N/A",
}
# Fields defaulted to "0": balance-sheet items (constant 0 for plain ETFs) + the
# fund-accounting/transfer-agent gains & flows (no feed — operator overrides in the sheet).
_ZERO_FIELDS = [
    "assetsAttrMiscSec", "assetsInvested",
    "amtPayOneYrBanksBorr", "amtPayOneYrCtrldComp", "amtPayOneYrOthAffil", "amtPayOneYrOther",
    "amtPayAftOneYrBanksBorr", "amtPayAftOneYrCtrldComp", "amtPayAftOneYrOthAffil", "amtPayAftOneYrOther",
    "delayDeliv", "standByCommit", "liquidPref",
    "netRealizedGainMon1", "netUnrealizedApprMon1", "netRealizedGainMon2",
    "netUnrealizedApprMon2", "netRealizedGainMon3", "netUnrealizedApprMon3",
    "mon1Sales", "mon1Redemption", "mon1Reinvestment",
    "mon2Sales", "mon2Redemption", "mon2Reinvestment",
    "mon3Sales", "mon3Redemption", "mon3Reinvestment",
]


def _fnum(x) -> float:
    try:
        return float(str(x).replace(",", "").strip())
    except (TypeError, ValueError):
        return 0.0


def _month_ranges(period: str) -> list[tuple[str, str]]:
    """The 3 reporting-period months as (YYYYMMDD start, YYYYMMDD end), chronological.

    For a period ending YYYY-MM, returns months MM-2, MM-1, MM (rtn1 earliest …
    rtn3 = the report month), matching the N-PORT monthlyTotReturn ordering.
    """
    y, m = int(period[:4]), int(period[5:7])
    out: list[tuple[str, str]] = []
    for back in (2, 1, 0):
        yy, mm = y, m - back
        while mm <= 0:
            mm += 12
            yy -= 1
        last = calendar.monthrange(yy, mm)[1]
        out.append((f"{yy:04d}{mm:02d}01", f"{yy:04d}{mm:02d}{last:02d}"))
    return out


def _return_formula(bbgid_cell: str, start: str, end: str) -> str:
    """Single-cell Bloomberg custom-total-return for one month (verified live)."""
    return (f'=BDP({bbgid_cell},"CUST_TRR_RETURN_HOLDING_PER",'
            f'"CUST_TRR_START_DT","{start}","CUST_TRR_END_DT","{end}","CUST_TRR_CRNCY","USD")')


def _signed_date(period: str) -> str:
    """A valid default dateSigned: last day of the month after the period end."""
    y, m = int(period[:4]), int(period[5:7])
    m += 1
    if m > 12:
        m, y = 1, y + 1
    return f"{y:04d}-{m:02d}-{calendar.monthrange(y, m)[1]:02d}"


def _clean_return(value: str) -> str:
    """A calculated return → 2dp string; blank/#N/A (no history) → schema-valid 'N/A'."""
    s = (value or "").strip()
    if not s or s.startswith("#"):
        return "N/A"
    try:
        return f"{float(s):.2f}"
    except ValueError:
        return "N/A"


# ── B.3 risk helpers ──────────────────────────────────────────


def _risk_formula(cell: str, mnemonic: str) -> str:
    """Bare single-field Bloomberg lookup on the bbgid cell (e.g. DUR_ADJ_MID)."""
    return f'=BDP({cell},"{mnemonic}")'


def _maybe_float(value) -> float | None:
    """Parse a calculated number; None for blank / #N/A (uncalculated off-terminal)."""
    s = str(value or "").replace(",", "").strip()
    if not s or s.startswith("#"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _maturity_years(maturity, period: str) -> float:
    """Years from the period end to a MATURITY date (ISO or M/D/YYYY); 0 if unparseable."""
    s = str(maturity or "").strip()
    if not s:
        return 0.0
    mat = None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            mat = datetime.strptime(s[:19] if " " in s else s[:10], fmt)
            break
        except ValueError:
            continue
    if mat is None:
        return 0.0
    end = datetime.strptime(_period_end_date(period), "%Y-%m-%d")
    return max(0.0, (mat - end).days / 365.25)


def _maturity_bucket(years: float) -> str:
    """Nearest N-PORT tenor bucket for a years-to-maturity."""
    for cut, bucket in _BUCKET_CUTS:
        if years < cut:
            return bucket
    return "30year"


def _is_investment_grade(rating) -> bool:
    """S&P IG (AAA…BBB-); blank/unrated → IG (treasuries carry no RTG_SP)."""
    s = str(rating or "").strip().upper()
    if not s or s.startswith("#") or s in {"NR", "N/A", "NA"}:
        return True
    return s in _IG_RATINGS


def _aggregate_risk(rows: list[dict], period: str) -> tuple[str, str, str]:
    """Aggregate per-debt-holding durations into B.3 JSON for one fund.

    Returns (cur_metrics_json, ig_json, nonig_json). All empty when no holding has a
    calculated duration (off-terminal) so the builder omits B.3 entirely.
    """
    dv01 = {b: 0.0 for b in _BUCKETS}
    dv100 = {b: 0.0 for b in _BUCKETS}
    ig = {b: 0.0 for b in _BUCKETS}
    nonig = {b: 0.0 for b in _BUCKETS}
    seen = False
    for r in rows:
        dur = _maybe_float(r.get("durAdj"))
        if dur is None:
            continue
        seen = True
        mv = _fnum(r.get("valUSD"))
        bucket = _maturity_bucket(_maturity_years(r.get("maturity"), period))
        dv01[bucket] += dur * mv * 1e-4
        dv100[bucket] += dur * mv * 1e-2
        spread = _maybe_float(r.get("spreadDur"))
        sdv01 = (spread if spread is not None else dur) * mv * 1e-4
        (ig if _is_investment_grade(r.get("ratingSP")) else nonig)[bucket] += sdv01
    if not seen:
        return "", "", ""
    cur_metric = {"curCd": "USD"}
    for b in _BUCKETS:
        cur_metric[f"dv01_{b}"] = f"{dv01[b]:.2f}"
        cur_metric[f"dv100_{b}"] = f"{dv100[b]:.2f}"
    ig_json = {b: f"{ig[b]:.2f}" for b in _BUCKETS}
    nonig_json = {b: f"{nonig[b]:.2f}" for b in _BUCKETS}
    return json.dumps([cur_metric]), json.dumps(ig_json), json.dumps(nonig_json)


# ── Build (custodian → filing master) ─────────────────────────


def _write_risk_sheet(wb, custodian_rows: list) -> None:
    """Add a 'risk' worksheet: one row per debt holding with Bloomberg duration =BDP cells."""
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet("risk")
    ws.append(RISK_HEADER)
    bbgid_col = get_column_letter(RISK_HEADER.index("bbgid") + 1)
    mnemonic = dict(_RISK_BDP_FIELDS)
    excel_row = 1
    for r in custodian_rows:
        ht = classify_holding(r)
        if ht is HoldingType.TREASURY:
            suffix = "Govt"
        elif ht is HoldingType.CORPORATE_BOND:
            suffix = "Corp"
        else:
            continue
        excel_row += 1
        cusip = (r.cusip or "").strip()
        bref = f"${bbgid_col}{excel_row}"
        literals = {
            "Account": r.account.upper(), "cusip": cusip, "bbgid": f"{cusip} {suffix}",
            "valUSD": f"{_fnum(r.market_value):.2f}",
        }
        ws.append([
            _risk_formula(bref, mnemonic[col]) if col in mnemonic else literals.get(col, "")
            for col in RISK_HEADER
        ])
    for er in range(2, ws.max_row + 1):
        for ci in range(1, len(RISK_HEADER) + 1):
            cell = ws.cell(row=er, column=ci)
            if not (isinstance(cell.value, str) and cell.value.startswith("=")):
                cell.number_format = "@"


def build_filing_master(
    custodian_rows: list, period: str, path: Path,
    ap_orders_path: Path | None = None,
) -> int:
    """Write the per-period filing master workbook. Returns the fund count.

    When ``ap_orders_path`` is given, monthly Sales/Redemption flows are aggregated
    from the AP creation/redemption order book and written as literal cells (the
    operator can still override). A ``risk`` worksheet of per-debt-holding Bloomberg
    duration ``=BDP`` formulas is always emitted for B.3 aggregation at split time.
    """
    from openpyxl.utils import get_column_letter

    end_date = _period_end_date(period)
    ranges = _month_ranges(period)
    signed = _signed_date(period)

    flows = flows_from_csv(Path(ap_orders_path), period) if ap_orders_path else {}

    by_acct: dict[str, list] = defaultdict(list)
    for r in custodian_rows:
        by_acct[r.account.upper()].append(r)

    rows: list[dict[str, str]] = []
    for acct in sorted(by_acct):
        rs = by_acct[acct]
        net = _fnum(rs[0].net_assets)
        liabs = sum(-_fnum(r.market_value) for r in rs if _fnum(r.market_value) < 0)
        rec = {c: "" for c in HEADER}
        rec["Account"] = acct
        rec["bbgid"] = f"{acct} US Equity"
        for c in _ZERO_FIELDS:
            rec[c] = "0"
        rec.update(_CONST)
        rec["repPdEnd"] = end_date
        rec["repPdDate"] = end_date
        rec["dateSigned"] = signed
        rec["netAssets"] = f"{net:.2f}"
        rec["totLiabs"] = f"{liabs:.2f}"
        rec["totAssets"] = f"{net + liabs:.2f}"
        # Capital flows from the AP order book (Sales/Redemption; reinvestment stays 0).
        for k, v in flows.get(acct, {}).items():
            rec[k] = v
        rows.append(rec)

    wb = Workbook()
    ws = wb.active
    ws.title = "filing"
    ws.append(HEADER)
    bbgid_col = get_column_letter(HEADER.index("bbgid") + 1)
    for idx, rec in enumerate(rows):
        excel_row = idx + 2
        bref = f"${bbgid_col}{excel_row}"
        out = []
        for col in HEADER:
            if col in RETURN_COLS:
                start, end = ranges[RETURN_COLS.index(col)]
                out.append(_return_formula(bref, start, end))
            else:
                out.append(rec.get(col, ""))
        ws.append(out)

    # Text-format the literal cells so Excel keeps dates/numbers exactly as written;
    # leave the return formula cells General so Bloomberg evaluates them.
    for excel_row in range(2, ws.max_row + 1):
        for col_idx in range(1, len(HEADER) + 1):
            cell = ws.cell(row=excel_row, column=col_idx)
            if not (isinstance(cell.value, str) and cell.value.startswith("=")):
                cell.number_format = "@"

    _write_risk_sheet(wb, custodian_rows)

    path.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp = tempfile.mkstemp(dir=path.parent, suffix=".xlsx")
    os.close(fd)
    try:
        wb.save(tmp)
        Path(tmp).replace(path)
    except Exception:
        Path(tmp).unlink(missing_ok=True)
        raise
    return len(rows)


def build_filing_master_from_custodian(
    custodian_path: Path, period: str, path: Path,
    ap_orders_path: Path | None = None,
) -> int:
    return build_filing_master(
        parse_custodian_csv(Path(custodian_path)), period, path, ap_orders_path)


# ── Read + split (filing master → filing_data.txt) ────────────


def read_filing_master(path: Path) -> list[dict[str, str]]:
    """Read the filing master into row dicts; returns resolve cached values (#N/A → N/A)."""
    wb = load_workbook(path, data_only=True)
    ws = wb["filing"] if "filing" in wb.sheetnames else wb.active
    raw_rows = list(ws.iter_rows(values_only=True))
    if not raw_rows:
        return []
    header = [_cell_to_str(h).strip() for h in raw_rows[0] if h is not None]
    out: list[dict[str, str]] = []
    for raw in raw_rows[1:]:
        if raw is None or all(c is None for c in raw):
            continue
        rec: dict[str, str] = {}
        for i, col in enumerate(header):
            v = _cell_to_str(raw[i]) if i < len(raw) else ""
            rec[col] = _clean_return(v) if col in RETURN_COLS else v
        out.append(rec)
    return out


def read_risk_sheet(path: Path) -> list[dict[str, str]]:
    """Read the 'risk' sheet into row dicts (cached duration values); [] if absent."""
    wb = load_workbook(path, data_only=True)
    if "risk" not in wb.sheetnames:
        return []
    raw_rows = list(wb["risk"].iter_rows(values_only=True))
    if not raw_rows:
        return []
    header = [_cell_to_str(h).strip() for h in raw_rows[0] if h is not None]
    out: list[dict[str, str]] = []
    for raw in raw_rows[1:]:
        if raw is None or all(c is None for c in raw):
            continue
        out.append({
            header[i]: (_cell_to_str(raw[i]) if i < len(raw) else "")
            for i in range(len(header))
        })
    return out


_SECTIONS = [
    ("Submission", ["submissionType", "liveTestFlag", "repPdEnd", "repPdDate",
                    "isFinalFiling", "dateSigned"]),
    ("Fund Financials", ["totAssets", "totLiabs", "netAssets"]),
    ("Balance Sheet Items", ["assetsAttrMiscSec", "assetsInvested",
                             "amtPayOneYrBanksBorr", "amtPayOneYrCtrldComp",
                             "amtPayOneYrOthAffil", "amtPayOneYrOther",
                             "amtPayAftOneYrBanksBorr", "amtPayAftOneYrCtrldComp",
                             "amtPayAftOneYrOthAffil", "amtPayAftOneYrOther",
                             "delayDeliv", "standByCommit", "liquidPref",
                             "isNonCashCollateral"]),
    ("Returns (rtn1-3 from Bloomberg; gains from fund accounting)",
     ["rtn1", "rtn2", "rtn3", "netRealizedGainMon1", "netUnrealizedApprMon1",
      "netRealizedGainMon2", "netUnrealizedApprMon2", "netRealizedGainMon3",
      "netUnrealizedApprMon3"]),
    ("Flows (from transfer agent / fund accounting)",
     ["mon1Sales", "mon1Redemption", "mon1Reinvestment",
      "mon2Sales", "mon2Redemption", "mon2Reinvestment",
      "mon3Sales", "mon3Redemption", "mon3Reinvestment"]),
    ("Designated Index", ["nameDesignatedIndex", "indexIdentifier"]),
]


def _format_filing_data(rec: dict[str, str], period: str) -> str:
    lines = [f"# {rec.get('Account', '')} {period} filing data",
             "# rtn1-3 from Bloomberg; net/total assets from custodian; gains/flows from fund accounting.",
             ""]
    for title, keys in _SECTIONS:
        lines.append(f"# {title}")
        for k in keys:
            lines.append(f"{k}={rec.get(k, '')}")
        lines.append("")
    if rec.get("curMetricsJson"):
        lines.append("# B.3 Risk Metrics (Bloomberg durations × custodian market value)")
        for k in ("curMetricsJson", "creditSprdRiskIgJson", "creditSprdRiskNonigJson"):
            lines.append(f"{k}={rec.get(k, '')}")
        lines.append("")
    return "\n".join(lines).rstrip() + "\n"


def split_filing_master(
    master_path: Path, funds_dir: Path, period: str, accounts: list[str] | None = None,
    dry_run: bool = False,
) -> list[tuple[str, Path]]:
    """Write each fund's filings/<period>/filing_data.txt from the filing master."""
    rows = read_filing_master(master_path)
    risk_by_acct: dict[str, list[dict]] = defaultdict(list)
    for rr in read_risk_sheet(master_path):
        risk_by_acct[(rr.get("Account") or "").strip().upper()].append(rr)
    target = {a.upper() for a in accounts} if accounts else None
    results: list[tuple[str, Path]] = []
    for rec in rows:
        acct = (rec.get("Account") or "").strip()
        if not acct or (target and acct not in target):
            continue
        cur_json, ig_json, nonig_json = _aggregate_risk(risk_by_acct.get(acct.upper(), []), period)
        if cur_json:
            rec["curMetricsJson"] = cur_json
            rec["creditSprdRiskIgJson"] = ig_json
            rec["creditSprdRiskNonigJson"] = nonig_json
        out_dir = funds_dir / acct.lower() / "filings" / period
        path = out_dir / "filing_data.txt"
        if not dry_run:
            out_dir.mkdir(parents=True, exist_ok=True)
            path.write_text(_format_filing_data(rec, period), encoding="utf-8")
        results.append((acct, path))
    return results
